"""
excel_parser.py - Excel and CSV Workbook Parser

Reads .xlsx and .csv workload tracker files.

Workload trackers are flat single-sheet files where each row is a feed,
and columns describe: feed name/ID, assignee, volume, frequency, time
estimate, and the resulting workload points.

Because the column headers vary by workbook (some say "Feed Name", others
"FeedID" or "feed_code"), this module uses the same canonical field alias
approach as flatfile_parser.py — so field resolution is consistent.

The scoring formula is NOT hardcoded here. It lives in:
    data/audit/workload/scoring_config.json

That file defines the weights, multipliers, and formula. Swap it out without
touching this file.

Supported formats:
    .xlsx  — via openpyxl (no pandas required)
    .csv   — via Python csv module (no extra dependencies)

If pandas is available it improves .xlsx handling but is not required.

Example workbook row (conceptual, columns vary):
    FeedName  | Assignee | State | Volume     | Frequency | TimeMins | Points
    mdjdws    | CK       | MD    | 5,000,000  | Weekly    | 90       | 5.0
    gadocws   | CK       | GA    | 10,000     | Weekly    | 20       | 1.0
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


# ---------------------------------------------------------------------------
# Canonical field name resolver (mirrors flatfile_parser._FIELD_ALIASES)
# ---------------------------------------------------------------------------

_FIELD_ALIASES: Dict[str, str] = {
    # Feed identity
    "feedname": "feed_name",
    "feed_name": "feed_name",
    "feedid": "feed_id",
    "feed_id": "feed_id",
    "feedcode": "feed_id",
    "feed_code": "feed_id",
    "feed": "feed_name",
    "feedlabel": "feed_name",
    "feed_label": "feed_name",
    "shortname": "feed_id",
    "short_name": "feed_id",
    "name": "feed_name",
    "projectfeedid": "pfid",
    "project_feed_id": "pfid",
    "pfid": "pfid",

    # State
    "state": "state",
    "statecode": "state",
    "state_code": "state",
    "stateabbr": "state",
    "st": "state",

    # Assignee
    "assignee": "assignee",
    "assigned_to": "assignee",
    "assignedto": "assignee",
    "analyst": "assignee",
    "qaperson": "assignee",
    "qa_person": "assignee",
    "owner": "assignee",
    "staff": "assignee",
    "person": "assignee",
    "analyst_name": "assignee",
    "username": "assignee",

    # Volume
    "volume": "volume",
    "recordcount": "volume",
    "record_count": "volume",
    "records": "volume",
    "avg_volume": "volume",
    "avgvolume": "volume",
    "avgrecords": "volume",
    "avg_records": "volume",
    "rowcount": "volume",
    "row_count": "volume",
    "volumeperdelivery": "volume",
    "volume_per_delivery": "volume",
    "volumeperweek": "volume",
    "volumepermonth": "volume",

    # Frequency
    "frequency": "frequency",
    "delivery_frequency": "frequency",
    "deliveryfrequency": "frequency",
    "cadence": "frequency",
    "schedule": "frequency",
    "freq": "frequency",
    "deliveryschedule": "frequency",
    "delivery_schedule": "frequency",

    # Time
    "time": "time_minutes",
    "timeminutes": "time_minutes",
    "time_minutes": "time_minutes",
    "timemins": "time_minutes",
    "time_mins": "time_minutes",
    "qatime": "time_minutes",
    "qa_time": "time_minutes",
    "estimatedtime": "time_minutes",
    "estimated_time": "time_minutes",
    "avgtime": "time_minutes",
    "avg_time": "time_minutes",
    "timehours": "time_hours",
    "time_hours": "time_hours",

    # Difficulty
    "difficulty": "difficulty",
    "complexitylevel": "difficulty",
    "complexity_level": "difficulty",
    "complexity": "difficulty",
    "difficultylevel": "difficulty",
    "difficulty_level": "difficulty",

    # Points / score
    "points": "workload_points",
    "workload_points": "workload_points",
    "workloadpoints": "workload_points",
    "workloadscore": "workload_points",
    "workload_score": "workload_points",
    "score": "workload_points",
    "weight": "workload_points",
    "weightedpoints": "workload_points",
    "weighted_points": "workload_points",
    "load": "workload_points",
    "loadpoints": "workload_points",

    # DA / analyst responsible — all common variations used in workload trackers
    "da_responsible":              "assignee",
    "daresponsible":               "assignee",
    "da_responsible_person":       "assignee",
    "state_sme":                   "assignee",
    "statesme":                    "assignee",
    "sme":                         "assignee",
    "primary_da":                  "assignee",
    "primaryda":                   "assignee",
    "responsible_da":              "assignee",
    "responsibleda":               "assignee",
    "da":                          "assignee",
    "lead":                        "assignee",
    "responsible":                 "assignee",
    "responsibility":              "assignee",
    "assigned_da":                 "assignee",
    "assignedda":                  "assignee",

    # Analyst / assigned analyst variants
    "assigned_analyst":            "assignee",
    "assignedanalyst":             "assignee",
    "performance_analyst":         "assignee",
    "performanceanalyst":          "assignee",
    "acquisition_analyst":         "assignee",
    "acquisitionanalyst":          "assignee",

    # Lead List Responsibility — separate role key (not the same as DA Responsible).
    # These entries carry higher workload weight; keeping them distinct prevents
    # the column-overwrite problem when a row has both LL and DA columns.
    "lead_list_responsibility":    "ll_assignee",
    "leadlistresponsibility":      "ll_assignee",
    "lead_list_da":                "ll_assignee",
    "leadlistda":                  "ll_assignee",
    "ll_responsibility":           "ll_assignee",
    "llresponsibility":            "ll_assignee",

    # DQS Responsible — separate role key
    "dqs_responsible":             "dqs_assignee",
    "dqsresponsible":              "dqs_assignee",
    "dqs":                         "dqs_assignee",

    # Back-up / backup DA — separate role key
    "back-up_da_responsible":      "backup_assignee",
    "backup_da_responsible":       "backup_assignee",
    "backupdaresponsible":         "backup_assignee",
    "back-updalresponsible":       "backup_assignee",
    "back-up_da":                  "backup_assignee",
    "backup_da":                   "backup_assignee",
    "backup_analyst":              "backup_assignee",
    "back-up_analyst":             "backup_assignee",

    # Volume — handle multi-line headers like 'Ave Volume \nWeekly/Monthly 2025'
    "ave_volume_weekly/monthly_2025": "volume",
    "ave_volume_weekly/monthly_2024": "volume",
    "ave_volume_weekly/monthly_2023": "volume",
    "ave_volume_weekly/monthly_2022": "volume",
    "ave_volume_weekly/monthly":    "volume",
    "avevolumeweeklymonthly2025":   "volume",
    "avevolumeweeklymonthly2024":   "volume",
    "avevolumeweeklymonthly2023":   "volume",
    "avevolumeweeklymonthly2022":   "volume",
    "avevolumeweeklymonthly":       "volume",
    "avevolume":                    "volume",
    "ave_volume":                   "volume",

    # Time
    "time_to_complete":            "time_minutes",
    "timetocomplete":              "time_minutes",
    "time_per_source":             "time_minutes",
    "timepersource":               "time_minutes",

    # Type/difficulty combos (e.g. "Type/Difficulty" column)
    "type_difficulty":             "difficulty",
    "typedifficulty":              "difficulty",
    "type_complexity":             "difficulty",
    "typecomplexity":              "difficulty",

    # SOP / notes
    "sop": "sop",
    "sopnotes": "sop",
    "sop_notes": "sop",
    "notes": "notes",
    "comments": "notes",
    "comment": "notes",

    # Status
    "status": "status",
    "feedstatus": "status",
    "feed_status": "status",
    "active": "is_active",
    "isactive": "is_active",
    "is_active": "is_active",

    # County/jurisdiction
    "county": "county",
    "jurisdiction": "county",
    "court": "county",
    "courtname": "county",
    "court_name": "county",

    # Data type
    "datatype": "data_type",
    "data_type": "data_type",
    "recordtype": "data_type",
    "record_type": "data_type",
    "type": "data_type",
    "category": "data_type",
}


def normalize_col(name: str) -> str:
    """Normalize a raw column header to a canonical field name."""
    # Strip whitespace including newlines/tabs that appear in multi-line Excel headers
    cleaned = name.strip().replace("\n", " ").replace("\r", " ").replace("\t", " ")
    # Collapse multiple spaces
    while "  " in cleaned:
        cleaned = cleaned.replace("  ", " ")
    # Try direct alias with underscores preserved
    result = _FIELD_ALIASES.get(cleaned.lower().strip().replace(" ", "_"), None)
    if result:
        return result
    # Try alias with all separators stripped
    key = cleaned.lower().replace(" ", "").replace("-", "").replace("_", "").replace("/", "")
    for alias_key, canonical in _FIELD_ALIASES.items():
        if alias_key.replace("_", "") == key:
            return canonical
    return cleaned.lower().strip().replace(" ", "_")


# ---------------------------------------------------------------------------
# Excel / CSV Parser
# ---------------------------------------------------------------------------

class WorkbookParser:
    """
    Reads a single-sheet .xlsx or .csv workload tracker.

    Normalizes column headers, returns a list of dicts with canonical keys.
    Handles:
    - Extra whitespace in headers and values
    - Numeric fields with commas (1,000,000 -> 1000000)
    - Empty rows (skipped)
    - BOM characters in CSV
    - Multiple encodings (UTF-8, latin-1)
    """

    def __init__(self, filepath: str, sheet_name: Optional[str] = None):
        self.filepath = Path(filepath)
        self.sheet_name = sheet_name  # For .xlsx: None = first sheet
        self.raw_headers: List[str] = []
        self.canonical_headers: List[str] = []
        self.records: List[Dict[str, Any]] = []
        self.parse_errors: List[Dict[str, Any]] = []
        self.row_count = 0

    def parse(self) -> "WorkbookParser":
        """Parse the file. Returns self for chaining."""
        suffix = self.filepath.suffix.lower()
        if suffix == ".csv":
            self._parse_csv()
        elif suffix in (".xlsx", ".xlsm", ".xltx"):
            self._parse_xlsx()
        elif suffix == ".xls":
            self.parse_errors.append({
                "error": ".xls format not supported. Please save as .xlsx or export as .csv.",
                "file": str(self.filepath),
            })
        else:
            self.parse_errors.append({
                "error": f"Unsupported file extension: {suffix}",
                "file": str(self.filepath),
            })
        return self

    def _parse_csv(self):
        """Parse a .csv file using the built-in csv module."""
        # Try UTF-8-BOM first (common Excel CSV export), fall back to latin-1
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(self.filepath, newline="", encoding=encoding) as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                break
            except UnicodeDecodeError:
                continue
        else:
            self.parse_errors.append({"error": "Could not decode CSV file.", "file": str(self.filepath)})
            return

        if not rows:
            self.parse_errors.append({"error": "CSV file is empty.", "file": str(self.filepath)})
            return

        self.raw_headers = [h.strip() for h in rows[0]]
        self.canonical_headers = [normalize_col(h) for h in self.raw_headers]

        for i, row in enumerate(rows[1:], start=2):
            if not any(cell.strip() for cell in row):
                continue  # skip blank rows
            while len(row) < len(self.canonical_headers):
                row.append("")
            record = {
                self.canonical_headers[j]: _coerce_value(row[j].strip())
                for j in range(len(self.canonical_headers))
            }
            record["_row"] = i
            self.records.append(record)
            self.row_count += 1

    def _parse_xlsx(self):
        """Parse a .xlsx file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            # Try pandas as fallback
            try:
                self._parse_xlsx_pandas()
                return
            except ImportError:
                self.parse_errors.append({
                    "error": (
                        "Neither openpyxl nor pandas is installed. "
                        "Install with: pip install openpyxl\n"
                        "Or export the workbook as .csv and use that instead."
                    ),
                    "file": str(self.filepath),
                })
                return

        wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)

        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                self.parse_errors.append({
                    "error": f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}",
                    "file": str(self.filepath),
                })
                return
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            self.parse_errors.append({"error": "Worksheet is empty.", "file": str(self.filepath)})
            return

        self.raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        self.canonical_headers = [normalize_col(h) for h in self.raw_headers]

        for i, row in enumerate(rows[1:], start=2):
            if not any(cell is not None for cell in row):
                continue  # skip blank rows
            row_list = list(row)
            while len(row_list) < len(self.canonical_headers):
                row_list.append(None)
            record = {
                self.canonical_headers[j]: _coerce_value(row_list[j])
                for j in range(len(self.canonical_headers))
            }
            record["_row"] = i
            self.records.append(record)
            self.row_count += 1

    def _parse_xlsx_pandas(self):
        """Fallback: parse .xlsx using pandas."""
        import pandas as pd
        kwargs = {"sheet_name": self.sheet_name or 0, "engine": "openpyxl"}
        df = pd.read_excel(self.filepath, **kwargs)
        self.raw_headers = list(df.columns.astype(str))
        self.canonical_headers = [normalize_col(h) for h in self.raw_headers]
        df.columns = self.canonical_headers

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            record = {k: _coerce_value(v) for k, v in row.items()}
            record["_row"] = i
            if any(v for k, v in record.items() if not k.startswith("_") and v):
                self.records.append(record)
                self.row_count += 1

    def available_sheets(self) -> List[str]:
        """Return sheet names for .xlsx files without fully parsing."""
        if self.filepath.suffix.lower() not in (".xlsx", ".xlsm"):
            return []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []

    # Sheets that look like config/reference tabs — skip during data scanning
    _SKIP_SHEET_KEYWORDS = (
        "param", "parameter", "weight", "config", "legend", "key",
        "readme", "notes", "instructions", "reference", "lookup", "summary",
        # Court/population reference sheets (not workload data)
        "nvscourt", "nvscourtpop", "court", "population", "_pop", "-pop",
        "state-county", "statecounty", "county_pop", "countypop",
        # Production tracking (not workload assignment)
        "production",
    )

    def _is_data_sheet(self, name: str) -> bool:
        """Return True if the sheet name looks like a data sheet (not a config tab)."""
        n = name.lower().strip()
        return not any(kw in n for kw in self._SKIP_SHEET_KEYWORDS)

    def parse_all_sheets(self, skip_non_data: bool = True) -> "WorkbookParser":
        """
        Parse every sheet in an .xlsx workbook, merging all rows into
        a single records list.  Each record gets a ``_sheet`` key with
        the sheet name so downstream code knows where the row came from.

        Sheets whose names match _SKIP_SHEET_KEYWORDS are skipped when
        skip_non_data=True (the default).  Force-include every sheet by
        passing skip_non_data=False.

        Falls back to single-sheet parse for .csv files.
        """
        if self.filepath.suffix.lower() not in (".xlsx", ".xlsm"):
            return self.parse()      # CSV — no multi-sheet

        try:
            import openpyxl
        except ImportError:
            return self.parse()

        wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        sheets_to_parse = [
            s for s in sheet_names
            if (not skip_non_data) or self._is_data_sheet(s)
        ]

        if not sheets_to_parse:
            # Nothing looks like data — fall back to parsing all sheets
            sheets_to_parse = sheet_names

        all_records: List[Dict[str, Any]] = []
        all_raw_headers: List[str] = []
        for sheet in sheets_to_parse:
            sub = WorkbookParser(str(self.filepath), sheet_name=sheet)
            sub.parse()
            if sub.parse_errors and not sub.records:
                self.parse_errors.extend(sub.parse_errors)
                continue
            for rec in sub.records:
                rec["_sheet"] = sheet
            all_records.extend(sub.records)
            # Collect all unique raw headers across sheets for diagnostics
            for h in sub.raw_headers:
                if h not in all_raw_headers:
                    all_raw_headers.append(h)

        self.records = all_records
        self.row_count = len(all_records)
        self.raw_headers = all_raw_headers
        self.canonical_headers = list({normalize_col(h) for h in all_raw_headers})
        return self

    def read_params_sheet(self) -> Dict[str, Any]:
        """
        Look for a parameters / weights sheet in the workbook and read it
        as a key→value config dict.

        Expected format (either orientation works):
            Column A: parameter name  (e.g. "volume_weight", "time_weight")
            Column B: value           (e.g. 1.0)

        OR a two-row layout:
            Row 1: parameter names
            Row 2: values

        Returns an empty dict if no params sheet is found.
        """
        if self.filepath.suffix.lower() not in (".xlsx", ".xlsm"):
            return {}
        try:
            import openpyxl
        except ImportError:
            return {}

        wb = openpyxl.load_workbook(self.filepath, data_only=True, read_only=True)
        sheet_names = wb.sheetnames

        _PARAM_KEYWORDS = ("param", "parameter", "weight", "config", "scoring")
        params_sheet = next(
            (s for s in sheet_names
             if any(kw in s.lower() for kw in _PARAM_KEYWORDS)),
            None,
        )

        if not params_sheet:
            wb.close()
            return {}

        ws = wb[params_sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            return {}

        params: Dict[str, Any] = {}

        # Detect layout: if first cell of row 1 is a string and first cell
        # of row 2 is a number → two-row header/value layout.
        if (
            len(rows) >= 2
            and isinstance(rows[0][0], str)
            and rows[1][0] is not None
            and not isinstance(rows[1][0], str)
        ):
            # Two-row layout
            headers = [str(h).strip() if h else "" for h in rows[0]]
            values  = rows[1]
            for h, v in zip(headers, values):
                if h and v is not None:
                    key = h.lower().strip().replace(" ", "_")
                    params[key] = _coerce_value(v)
        else:
            # Column A/B layout (most common)
            for row in rows:
                if len(row) < 2:
                    continue
                k, v = row[0], row[1]
                if k is None or v is None:
                    continue
                key = str(k).lower().strip().replace(" ", "_")
                if key and key not in ("parameter", "param", "key", "name", "setting"):
                    params[key] = _coerce_value(v)

        return params

    def summary(self) -> Dict[str, Any]:
        return {
            "filepath": str(self.filepath),
            "filename": self.filepath.name,
            "row_count": self.row_count,
            "fields_detected": self.canonical_headers,
            "raw_headers": self.raw_headers,
            "parse_errors": self.parse_errors,
        }


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _coerce_value(val: Any) -> Any:
    """
    Normalize a cell value:
    - Strip whitespace from strings
    - Remove comma-separators from numbers (1,000,000 -> 1000000.0)
    - Convert Excel numeric dates back to strings if they look like dates
    - Pass through None as ""
    """
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, bool):
        return val
    s = str(val).strip()
    # Try stripping commas for large numbers
    stripped = s.replace(",", "")
    if stripped and stripped.replace(".", "").replace("-", "").isdigit():
        try:
            return float(stripped) if "." in stripped else int(stripped)
        except ValueError:
            pass
    return s
